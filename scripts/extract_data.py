"""
数据提取脚本：从 Excel 提取校园环线时刻表数据，生成 JSON 供 PWA 使用。

输入: 国防科大校园环线时刻表_秒级可校准版 (1).xlsx
输出: data/departures.json, data/route_params.json, data/stations.json, data/arrival_predictions.json
"""

import openpyxl
import json
import os
import math
from collections import OrderedDict

EXCEL_FILE = "国防科大校园环线时刻表_秒级可校准版 (1).xlsx"
OUTPUT_DIR = "public/data"
WAYPOINT_FILE = "站点经纬度信息路径细化.txt"

# ─── 物理模型参数 ───
SPEED_KMH = 25.0          # 校园限速 km/h
SPEED_MS = SPEED_KMH / 3.6  # 6.944 m/s
DWELL_SECONDS = 15        # 每站上下客停留（校园环线人少）
ACCEL_DECEL_PENALTY = 8   # 启停损耗：25km/h→0 减速度约1.5m/s²约需4.6s，加速同理，合计约8s


def haversine(lng1: float, lat1: float, lng2: float, lat2: float) -> float:
    """计算两点间距离（米）"""
    R = 6371000
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlng / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

# ─── WGS-84 → GCJ-02 坐标转换 ───
# 高德地图使用 GCJ-02（火星坐标系），谷歌地图使用 WGS-84
# 从谷歌地图获取的坐标需转换后才能在高德地图准确显示

def wgs84_to_gcj02(lng: float, lat: float) -> tuple[float, float]:
    """WGS-84 转 GCJ-02"""
    a = 6378245.0
    ee = 0.006693421622965943

    def _transform_lat(x: float, y: float) -> float:
        ret = -100.0 + 2.0 * x + 3.0 * y + 0.2 * y * y + 0.1 * x * y + 0.2 * math.sqrt(abs(x))
        ret += (20.0 * math.sin(6.0 * x * math.pi) + 20.0 * math.sin(2.0 * x * math.pi)) * 2.0 / 3.0
        ret += (20.0 * math.sin(y * math.pi) + 40.0 * math.sin(y / 3.0 * math.pi)) * 2.0 / 3.0
        ret += (160.0 * math.sin(y / 12.0 * math.pi) + 320.0 * math.sin(y * math.pi / 30.0)) * 2.0 / 3.0
        return ret

    def _transform_lng(x: float, y: float) -> float:
        ret = 300.0 + x + 2.0 * y + 0.1 * x * x + 0.1 * x * y + 0.1 * math.sqrt(abs(x))
        ret += (20.0 * math.sin(6.0 * x * math.pi) + 20.0 * math.sin(2.0 * x * math.pi)) * 2.0 / 3.0
        ret += (20.0 * math.sin(x * math.pi) + 40.0 * math.sin(x / 3.0 * math.pi)) * 2.0 / 3.0
        ret += (150.0 * math.sin(x / 12.0 * math.pi) + 300.0 * math.sin(x / 30.0 * math.pi)) * 2.0 / 3.0
        return ret

    dlat = _transform_lat(lng - 105.0, lat - 35.0)
    dlng = _transform_lng(lng - 105.0, lat - 35.0)
    radlat = lat / 180.0 * math.pi
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * math.pi)
    dlng = (dlng * 180.0) / (a / sqrtmagic * math.cos(radlat) * math.pi)
    return lng + dlng, lat + dlat


# ─── 站点坐标（谷歌地图 WGS-84 → 自动转为 GCJ-02）───

_WGS84_COORDS = {
    "研究生宿舍楼": (113.045661, 28.257994),
    "东门":         (113.047749, 28.260392),
    "2号宿舍楼":    (113.045821, 28.262810),
    "军体活动中心":  (113.042703, 28.264276),
    "激光所":       (113.042397, 28.269064),
    "超算中心":      (113.041494, 28.271481),
    "北门":         (113.038932, 28.272043),
    "高超楼":       (113.036590, 28.271516),
    "理学院":       (113.036611, 28.268852),
    "二食堂":       (113.037879, 28.260412),
    "5号宿舍楼":    (113.039804, 28.258361),
    "305教学楼":    (113.043036, 28.257571),
    "一食堂":       (113.044985, 28.259916),
    "门诊部":       (113.047260, 28.260459),
    "1号宿舍楼":    (113.044867, 28.263323),
    "水上训练中心":  (113.039453, 28.262648),
    "教勤连":       (113.041696, 28.265302),
    "图书馆":       (113.044942, 28.260047),
    "二食堂北":      (113.037881, 28.260462),
}

# 转换为 GCJ-02
STATION_COORDS = {}
for name, wgs in _WGS84_COORDS.items():
    gcj_lng, gcj_lat = wgs84_to_gcj02(wgs[0], wgs[1])
    STATION_COORDS[name] = (gcj_lat, gcj_lng)


def parse_departures(wb):
    """读取 <<结构化数据>> 工作表，生成 departures 列表"""
    ws = wb["结构化数据"]
    departures = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue

        record_id = str(row[0]).strip()
        date_type_raw = str(row[1]).strip() if row[1] else ""
        route = str(row[2]).strip() if row[2] else ""
        original_label = str(row[3]).strip() if row[3] else ""
        shift_name = str(row[4]).strip() if row[4] else ""
        trip_seq = int(row[5]) if row[5] else 1
        departure_time = str(row[6]).strip() if row[6] else ""
        departure_time_hhmm = str(row[7]).strip() if row[7] else ""
        departure_minutes = int(row[8]) if row[8] else 0
        departure_station = str(row[9]).strip() if row[9] else ""
        is_gaochao_raw = str(row[10]).strip() if row[10] else "否"
        driver = str(row[11]).strip() if row[11] else ""
        vehicle_no = str(row[12]).strip() if row[12] else ""
        data_status = str(row[13]).strip() if row[13] else ""
        remark = str(row[14]).strip() if row[14] else ""
        source = str(row[15]).strip() if row[15] else ""

        # Date type mapping
        if "工作日" in date_type_raw and ("周末" in date_type_raw or "节假日" in date_type_raw):
            date_type = "weekend_holiday"
        elif "周末" in date_type_raw or "节假日" in date_type_raw:
            date_type = "weekend_holiday"
        else:
            date_type = "weekday"

        # is_gaochao
        is_gaochao = is_gaochao_raw == "是"

        # Route key logic
        if route == "环线1路":
            route_key = "HX1_DINING" if is_gaochao else "HX1_NORMAL"
        elif route == "环线2路":
            route_key = "HX2_NORMAL"
        elif route == "环线3路":
            route_key = "HX3_GAOCHAO" if is_gaochao else "HX3_NORMAL"
        else:
            route_key = "UNKNOWN"

        # Pattern name
        pattern_names = {
            "HX1_NORMAL": "环线1路普通路线（研究生宿舍楼发车）",
            "HX1_DINING": "环线1路就餐专线（高超楼发车）",
            "HX2_NORMAL": "环线2路普通路线（研究生宿舍楼发车）",
            "HX3_NORMAL": "环线3路普通路线（研究生宿舍楼发车）",
            "HX3_GAOCHAO": "环线3路高超楼发车路线",
        }

        # Departure station
        if "高超楼" in departure_station or "系统" in departure_station:
            dep_station = "高超楼"
        else:
            dep_station = "研究生宿舍楼"

        # Confidence
        confidence = "speculative" if "推测" in data_status else "confirmed"

        departures.append({
            "recordId": record_id,
            "dateType": date_type,
            "route": route,
            "shiftName": shift_name,
            "tripSeq": trip_seq,
            "departureTime": departure_time,
            "departureMinutes": departure_minutes,
            "departureStation": dep_station,
            "isGaochaoDeparture": is_gaochao,
            "routeKey": route_key,
            "patternName": pattern_names.get(route_key, ""),
            "driver": driver if driver not in ("未提供", "") else "",
            "vehicleNo": vehicle_no if vehicle_no not in ("未提供", "") else "",
            "confidence": confidence,
            "remark": remark if remark and remark != "None" else "",
        })

    return departures


def parse_route_params(wb):
    """读取 <<秒级路线参数_可调整>> 工作表，按 route_key 分组"""
    ws = wb["秒级路线参数_可调整"]
    patterns = OrderedDict()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue

        route_key = str(row[1]).strip() if row[1] else ""
        stop_seq = int(row[4]) if row[4] else 0
        prev_stop = str(row[5]).strip() if row[5] else ""
        current_stop = str(row[6]).strip() if row[6] else ""
        distance_km = float(row[7]) if row[7] else 0
        base_seconds = int(row[8]) if row[8] else 0
        manual_seconds_raw = row[12]
        is_departure = str(row[17]).strip() if row[17] else ""
        is_return = str(row[18]).strip() if row[18] else ""

        manual_seconds = None
        if manual_seconds_raw is not None:
            try:
                manual_seconds = int(manual_seconds_raw)
            except (ValueError, TypeError):
                manual_seconds = None

        final_seconds = manual_seconds if manual_seconds is not None else base_seconds

        if route_key not in patterns:
            patterns[route_key] = {
                "routeKey": route_key,
                "route": "",
                "patternName": "",
                "departureStation": "",
                "totalSeconds": 0,
                "stopCount": 0,
                "stops": [],
            }

        if route_key.startswith("HX1"):
            route_name = "环线1路"
        elif route_key.startswith("HX2"):
            route_name = "环线2路"
        else:
            route_name = "环线3路"

        patterns[route_key]["route"] = route_name

        stop = {
            "routeStopKey": f"{route_key}|{stop_seq}",
            "routeKey": route_key,
            "stopSeq": stop_seq,
            "prevStop": prev_stop,
            "currentStop": current_stop,
            "distanceKm": distance_km,
            "baseSegmentSeconds": base_seconds,
            "manualSegmentSeconds": manual_seconds,
            "finalSegmentSeconds": final_seconds,
            "cumulativeSeconds": 0,
            "isDepartureStop": is_departure == "是",
            "isReturnStop": is_return == "是",
        }
        patterns[route_key]["stops"].append(stop)

    # Fix: HX2_NORMAL 军体活动中心 和 水上训练中心 顺序对调
    # 正确顺序: 1号宿舍楼→军体活动中心→水上训练中心→二食堂
    if "HX2_NORMAL" in patterns:
        stops = patterns["HX2_NORMAL"]["stops"]
        # 只交换站名，保持各位置的站间秒数不变
        for s in stops:
            if s["currentStop"] == "水上训练中心":
                s["currentStop"] = "军体活动中心"
            elif s["currentStop"] == "军体活动中心":
                s["currentStop"] = "水上训练中心"
        # 修正 prevStop 链
        for idx in range(1, len(stops)):
            stops[idx]["prevStop"] = stops[idx - 1]["currentStop"]
        # 修正 routeStopKey
        for s in stops:
            s["routeStopKey"] = f'{s["routeKey"]}|{s["stopSeq"]}'

    # Recalculate cumulative seconds
    for rk, pattern in patterns.items():
        cum = 0
        departure_station = ""
        for stop in pattern["stops"]:
            if stop["isDepartureStop"]:
                departure_station = stop["currentStop"]
            cum += stop["finalSegmentSeconds"]
            stop["cumulativeSeconds"] = cum

        pattern["totalSeconds"] = cum
        pattern["stopCount"] = len(pattern["stops"])
        pattern["departureStation"] = departure_station

        pattern_names = {
            "HX1_NORMAL": "环线1路普通路线（研究生宿舍楼发车）",
            "HX1_DINING": "环线1路就餐专线（高超楼发车）",
            "HX2_NORMAL": "环线2路普通路线（研究生宿舍楼发车）",
            "HX3_NORMAL": "环线3路普通路线（研究生宿舍楼发车）",
            "HX3_GAOCHAO": "环线3路高超楼发车路线",
        }
        pattern["patternName"] = pattern_names.get(rk, "")

    return list(patterns.values())


def compute_arrival_predictions(departures, route_patterns):
    """预计算每条发车记录在每站的预计到站时间"""
    pattern_map = {p["routeKey"]: p for p in route_patterns}
    predictions = []

    for dep in departures:
        rk = dep["routeKey"]
        pattern = pattern_map.get(rk)
        if not pattern:
            continue

        for stop in pattern["stops"]:
            arrival_minutes = dep["departureMinutes"] + stop["cumulativeSeconds"] / 60.0
            display_minutes = arrival_minutes % 1440
            # 精确到秒
            total_secs = int(dep["departureMinutes"] * 60 + stop["cumulativeSeconds"])
            ds = total_secs % 86400
            h = ds // 3600
            m = (ds % 3600) // 60
            s = ds % 60
            arrival_time = f"{h:02d}:{m:02d}:{s:02d}"

            predictions.append({
                "departureId": dep["recordId"],
                "dateType": dep["dateType"],
                "route": dep["route"],
                "shiftName": dep["shiftName"],
                "tripSeq": dep["tripSeq"],
                "departureTime": dep["departureTime"],
                "departureMinutes": dep["departureMinutes"],
                "departureStation": dep["departureStation"],
                "routeKey": rk,
                "stopSeq": stop["stopSeq"],
                "stopName": stop["currentStop"],
                "cumulativeSeconds": stop["cumulativeSeconds"],
                "arrivalMinutes": arrival_minutes,
                "arrivalTime": arrival_time,
                "isDepartureStop": stop["isDepartureStop"],
                "isReturnStop": stop["isReturnStop"],
                "driver": dep["driver"],
                "vehicleNo": dep["vehicleNo"],
                "confidence": dep["confidence"],
            })

    return predictions


def parse_stations(wb):
    """读取 <<站点字典>> 并合并 GPS 坐标"""
    ws = wb["站点字典"]
    stations = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue

        name = str(row[0]).strip()
        service_routes_str = str(row[1]).strip() if row[1] else ""
        location_note = str(row[2]).strip() if row[2] else ""
        remark = str(row[3]).strip() if row[3] and str(row[3]) != "None" else ""

        coords = STATION_COORDS.get(name, (28.257994, 113.045661))

        stations.append({
            "name": name,
            "serviceRoutes": service_routes_str,
            "locationNote": location_note,
            "remark": remark,
            "lat": round(coords[0], 6),
            "lng": round(coords[1], 6),
        })

    return stations


def parse_waypoints(filepath: str) -> dict[str, list[tuple[float, float]]]:
    """解析途经点文件，返回 路线名 → [(lng, lat), ...] 的映射（GCJ-02 坐标）"""
    if not os.path.exists(filepath):
        print(f"  途经点文件未找到: {filepath}，跳过路线细化")
        return {}

    route_map: dict[str, list[tuple[float, float]]] = {}
    current_route = ""
    current_path: list[tuple[float, float]] = []

    route_names = {
        "环线1路": "HX1_NORMAL",
        "环线2路": "HX2_NORMAL",
        "环线3路": "HX3_NORMAL",
        "就餐专线": "HX1_DINING",
    }

    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            if line.startswith(">"):
                # Waypoint: "> 28.xxx, 113.xxx"
                try:
                    parts = line.lstrip("> ").replace(" ", "").split(",")
                    lat, lng = float(parts[0]), float(parts[1])
                    gcj_lng, gcj_lat = wgs84_to_gcj02(lng, lat)
                    current_path.append((gcj_lng, gcj_lat))
                except (ValueError, IndexError):
                    continue
            elif "途径站点" in line:
                # Route header: "环线1路途径站点"
                for cn_name, key in route_names.items():
                    if cn_name in line:
                        # Save previous route
                        if current_route and current_path:
                            route_map[current_route] = current_path
                        current_route = key
                        current_path = []
                        break
            elif "," in line and any(c.isdigit() for c in line.split(",")[0]):
                # Stop coordinate: "站点名 28.xxx, 113.xxx"
                try:
                    # Extract all numeric parts after the station name
                    parts = line.split()
                    # Find the coordinate segment: "28.xxx," and "113.xxx"
                    lat_str = ""
                    lng_str = ""
                    for p in parts:
                        if "," in p:
                            lat_str = p.rstrip(",")
                        elif lat_str and p.replace(".", "").isdigit():
                            lng_str = p
                    if lat_str and lng_str:
                        lat, lng = float(lat_str), float(lng_str)
                        gcj_lng, gcj_lat = wgs84_to_gcj02(lng, lat)
                        current_path.append((gcj_lng, gcj_lat))
                except (ValueError, IndexError):
                    continue

    # Save last route
    if current_route and current_path:
        route_map[current_route] = current_path

    # Visually close all loop routes by connecting last point to first
    for rk, path in route_map.items():
        if len(path) >= 2:
            path.append(path[0])

    # Generate HX3_GAOCHAO: same physical route as HX3_NORMAL but starts at 高超楼
    if "HX3_NORMAL" in route_map and "HX3_GAOCHAO" not in route_map:
        hx3_path = route_map["HX3_NORMAL"]
        # Find 高超楼 in the path (approximately) and slice from there
        gaochao_coord = STATION_COORDS.get("高超楼", None)
        if gaochao_coord:
            # Find closest point to 高超楼
            best_idx = 0
            best_dist = float("inf")
            for i, (lng, lat) in enumerate(hx3_path):
                d = (lng - gaochao_coord[1])**2 + (lat - gaochao_coord[0])**2
                if d < best_dist:
                    best_dist = d
                    best_idx = i
            # HX3_GAOCHAO starts from 高超楼 through the rest of the loop
            # Note: HX3 goes 研究生→...→高超楼→理学院→...→研究生
            # HX3_GAOCHAO should go 高超楼→理学院→...→研究生→一食堂→...→高超楼
            # This means: from 高超楼 index to end, then from beginning to 高超楼 index
            route_map["HX3_GAOCHAO"] = hx3_path[best_idx:] + hx3_path[1:best_idx+1]

    return route_map


def parse_route_stops(filepath: str) -> dict[str, list[dict]]:
    """解析途经点文件中每条路线的站点名和坐标（GCJ-02），返回 routeKey → [{name, lng, lat}, ...]"""
    if not os.path.exists(filepath):
        return {}

    route_names = {
        "环线1路": "HX1_NORMAL",
        "环线2路": "HX2_NORMAL",
        "环线3路": "HX3_NORMAL",
        "就餐专线": "HX1_DINING",
    }

    route_stops: dict[str, list[dict]] = {}
    current_key = ""
    seen_in_route: set[str] = set()

    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith(">"):
                continue

            # Route header
            for cn_name, key in route_names.items():
                if cn_name in line and "途径站点" in line:
                    current_key = key
                    seen_in_route = set()
                    route_stops.setdefault(current_key, [])
                    break
            else:
                # Station line: "站点名 28.xxx, 113.xxx"
                if current_key and "," in line:
                    parts = line.split()
                    lat_str = ""
                    lng_str = ""
                    for p in parts:
                        if "," in p:
                            lat_str = p.rstrip(",")
                        elif lat_str and p.replace(".", "").isdigit():
                            lng_str = p
                    if lat_str and lng_str:
                        try:
                            lat, lng = float(lat_str), float(lng_str)
                            gcj_lng, gcj_lat = wgs84_to_gcj02(lng, lat)
                            # Extract station name (everything before the coordinates)
                            name = line[:line.index(lat_str)].rstrip()
                            if name not in seen_in_route:
                                seen_in_route.add(name)
                                route_stops[current_key].append({
                                    "name": name,
                                    "lng": round(gcj_lng, 6),
                                    "lat": round(gcj_lat, 6),
                                })
                        except (ValueError, IndexError):
                            continue

    # HX3_GAOCHAO: same stops as HX3, starting from 高超楼
    if "HX3_NORMAL" in route_stops:
        hx3_stops = route_stops["HX3_NORMAL"]
        # Find 高超楼 index
        gaochao_idx = 0
        for i, s in enumerate(hx3_stops):
            if "高超楼" in s["name"]:
                gaochao_idx = i
                break
        route_stops["HX3_GAOCHAO"] = hx3_stops[gaochao_idx:] + hx3_stops[1:gaochao_idx + 1]

    return route_stops


def compute_accurate_timings(
    route_paths: dict, route_stops: dict, route_params: list
) -> list:
    """
    基于路径坐标和物理模型重新计算到站秒数。
    返回更新后的 route_params 列表。
    """
    SPEED = SPEED_MS           # m/s
    DWELL = DWELL_SECONDS       # 秒
    PENALTY = ACCEL_DECEL_PENALTY  # 秒

    for rp in route_params:
        rk = rp["routeKey"]
        # HX3_GAOCHAO 与 HX3_NORMAL 走同一条路，直接复用
        if rk == "HX3_GAOCHAO":
            hx3_normal = next((x for x in route_params if x["routeKey"] == "HX3_NORMAL"), None)
            if hx3_normal:
                rp["stops"] = hx3_normal["stops"]
                rp["totalSeconds"] = hx3_normal["totalSeconds"]
            continue
        stops = route_stops.get(rk, [])
        path = route_paths.get(rk, [])

        if len(path) < 2 or len(stops) < 2:
            continue

        # 计算路径上每一段的距离（米）
        seg_distances = []
        for i in range(len(path) - 1):
            d = haversine(path[i][0], path[i][1], path[i + 1][0], path[i + 1][1])
            seg_distances.append(d)

        total_distance = sum(seg_distances)
        total_travel_seconds = total_distance / SPEED
        stop_count = len(stops)
        total_dwell = (stop_count - 1) * DWELL  # 发车站不停留
        total_penalty = stop_count * PENALTY
        total_seconds = total_travel_seconds + total_dwell + total_penalty

        # 为每个 stop 找到路径上最近的点，计算累计距离
        stop_cum_dist = []
        for stop in stops:
            # 找路径上距此 stop 最近的点
            best_dist = float("inf")
            best_idx = 0
            for i, (lng, lat) in enumerate(path):
                d = (lng - stop["lng"])**2 + (lat - stop["lat"])**2
                if d < best_dist:
                    best_dist = d
                    best_idx = i
            # 累计距离 = 从路径起点到 best_idx 的所有段距离之和
            cum = sum(seg_distances[:best_idx]) if best_idx > 0 else 0
            stop_cum_dist.append((stop["name"], cum, best_idx))

        # 将累计距离转换为秒数（包含停站和加减速）
        stop_seconds = []
        prev_idx = 0
        for idx, (name, cum_dist, path_idx) in enumerate(stop_cum_dist):
            if idx == 0:
                stop_seconds.append(0)  # 发车
            else:
                # 从前一站到当前站的距离
                seg_dist = cum_dist - stop_cum_dist[idx - 1][1]
                seg_time = seg_dist / SPEED
                seg_total = seg_time + DWELL + PENALTY
                stop_seconds.append(stop_seconds[-1] + seg_total)

        # 更新 route_params 中的 stops
        param_stops = rp["stops"]
        for i, s in enumerate(param_stops):
            if i < len(stop_seconds):
                new_cum = round(stop_seconds[i])
                s["finalSegmentSeconds"] = new_cum - (round(stop_seconds[i - 1]) if i > 0 else 0)
                s["cumulativeSeconds"] = new_cum
        rp["totalSeconds"] = round(stop_seconds[-1]) if stop_seconds else rp["totalSeconds"]

        print(f"  {rk}: {total_distance:.0f}m, travel={total_travel_seconds:.0f}s, "
              f"dwell={total_dwell:.0f}s, accel={total_penalty:.0f}s → {rp['totalSeconds']}s total")

    return route_params


def main():
    print(f"Reading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("Extracting departures...")
    departures = parse_departures(wb)
    with open(os.path.join(OUTPUT_DIR, "departures.json"), "w", encoding="utf-8") as f:
        json.dump(departures, f, ensure_ascii=False, indent=2)
    print(f"  -> {len(departures)} records")

    print("Extracting route params...")
    route_params = parse_route_params(wb)
    with open(os.path.join(OUTPUT_DIR, "route_params.json"), "w", encoding="utf-8") as f:
        json.dump(route_params, f, ensure_ascii=False, indent=2)
    print(f"  -> {len(route_params)} route patterns")

    print("Extracting stations...")
    stations = parse_stations(wb)
    with open(os.path.join(OUTPUT_DIR, "stations.json"), "w", encoding="utf-8") as f:
        json.dump(stations, f, ensure_ascii=False, indent=2)
    print(f"  -> {len(stations)} stations (WGS-84 converted to GCJ-02)")

    print("Generating route paths...")
    route_paths = parse_waypoints(WAYPOINT_FILE)
    for rk, pts in route_paths.items():
        print(f"  {rk}: {len(pts)} points")

    print("Generating route stops...")
    route_stops = parse_route_stops(WAYPOINT_FILE)
    for rk, stops in route_stops.items():
        print(f"  {rk}: {len(stops)} stops")

    print("Computing accurate timings from path distances...")
    route_params = compute_accurate_timings(route_paths, route_stops, route_params)

    # Save outputs
    with open(os.path.join(OUTPUT_DIR, "route_params.json"), "w", encoding="utf-8") as f:
        json.dump(route_params, f, ensure_ascii=False, indent=2)
    with open(os.path.join(OUTPUT_DIR, "route_paths.json"), "w", encoding="utf-8") as f:
        json.dump(route_paths, f, ensure_ascii=False)
    with open(os.path.join(OUTPUT_DIR, "route_stops.json"), "w", encoding="utf-8") as f:
        json.dump(route_stops, f, ensure_ascii=False)

    print("Computing arrival predictions...")
    predictions = compute_arrival_predictions(departures, route_params)
    with open(os.path.join(OUTPUT_DIR, "arrival_predictions.json"), "w", encoding="utf-8") as f:
        json.dump(predictions, f, ensure_ascii=False, indent=2)
    print(f"  -> {len(predictions)} prediction records")

    total_kb = sum(
        os.path.getsize(os.path.join(OUTPUT_DIR, f))
        for f in os.listdir(OUTPUT_DIR)
        if f.endswith(".json")
    ) / 1024
    print(f"\nDone! Total JSON size: {total_kb:.1f} KB")


if __name__ == "__main__":
    main()
